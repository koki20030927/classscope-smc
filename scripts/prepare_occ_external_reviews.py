#!/usr/bin/env python3
"""Normalize and match an OCC external Review workbook without modifying it."""

from __future__ import annotations

import argparse
import csv
import hashlib
import json
import re
import unicodedata
from collections import Counter, defaultdict
from pathlib import Path
from typing import Any, Iterable


OCC_COURSES = Path("data/occ/processed/occ_courses.csv")
OCC_PROFESSORS = Path("data/occ/processed/occ_professors.csv")
OCC_RELATIONSHIPS = Path("data/occ/processed/occ_course_professors.csv")
SEMESTERS = {"fall", "spring", "summer", "winter"}
CLASS_FORMATS = {"in_person", "online", "hybrid"}
RATING_FIELDS = ("professor_quality", "easy_a", "course_quality", "recommendation")
OUTPUT_FIELDS = (
    "source_type",
    "source_row_key",
    "source_row_number",
    "status",
    "issues",
    "course_code",
    "professor_name",
    "professor_quality",
    "easy_a",
    "course_quality",
    "recommendation",
    "class_format",
    "year_taken",
    "semester",
    "content",
)
SENSITIVE_HEADER = re.compile(
    r"(?:e-?mail|phone|mobile|student\s*id|address|account|private\s*note)",
    re.IGNORECASE,
)
EMAIL = re.compile(r"\b[A-Z0-9._%+-]+@[A-Z0-9.-]+\.[A-Z]{2,}\b", re.IGNORECASE)
PHONE = re.compile(r"(?<!\d)(?:\+?1[\s.-]?)?\(?\d{3}\)?[\s.-]\d{3}[\s.-]\d{4}(?!\d)")
STUDENT_ID = re.compile(r"\b(?:student\s*id|sid|id\s*number)\s*[:#-]?\s*[A-Z0-9-]{5,}\b", re.IGNORECASE)


def text(value: Any) -> str:
    if value is None:
        return ""
    return unicodedata.normalize("NFKC", str(value)).strip()


def compact(value: str) -> str:
    return re.sub(r"\s+", " ", text(value))


def normalize_course(value: str) -> str:
    value = compact(value).upper().replace("–", "-").replace("—", "-")
    value = re.sub(r"[-_/]+", " ", value)
    match = re.fullmatch(r"([A-Z&]+)\s*([A-Z]?)\s*(\d+[A-Z]?)", value)
    if not match:
        return compact(value)
    subject, prefix, number = match.groups()
    return f"{subject} {prefix}{number}"


def normalize_professor(value: str) -> str:
    value = compact(value).casefold()
    value = re.sub(r"^(?:prof(?:essor)?|dr)\.?\s+", "", value)
    value = re.sub(r"\b(?:jr|sr|ii|iii|iv)\.?$", "", value).strip()
    value = re.sub(r"[^\w\s]", " ", value)
    return re.sub(r"\s+", " ", value).strip()


def professor_variants(value: str) -> set[str]:
    variants = {normalize_professor(value)}
    if "," in value:
        last, first = value.split(",", 1)
        variants.add(normalize_professor(f"{first} {last}"))
    return {variant for variant in variants if variant}


def parse_rating(value: Any, *, required: bool) -> tuple[str, bool]:
    raw = text(value)
    if not raw:
        return "", not required
    try:
        numeric = float(raw)
    except ValueError:
        return "", False
    integer = int(numeric)
    if numeric != integer or integer < 1 or integer > 5:
        return "", False
    return str(integer), True


def parse_year(value: Any) -> tuple[str, bool]:
    raw = text(value)
    combined_term = re.fullmatch(r"(20\d{2})\s+(?:fall|spring|summer|winter)", raw, re.IGNORECASE)
    if combined_term:
        return combined_term.group(1), True
    try:
        numeric = float(raw)
    except ValueError:
        return "", False
    year = int(numeric)
    return (str(year), True) if numeric == year and 2000 <= year <= 2100 else ("", False)


def contains_personal_data(values: Iterable[str]) -> bool:
    joined = "\n".join(values)
    return bool(EMAIL.search(joined) or PHONE.search(joined) or STUDENT_ID.search(joined))


def read_rows(
    path: Path,
    *,
    sheet_selection: str,
    sheet_name: str,
    header_row: int,
) -> tuple[list[str], list[dict[str, Any]], dict[str, Any]]:
    if path.suffix.lower() == ".csv":
        if sheet_selection == "first_sheet":
            raise ValueError("first_sheet selection requires an .xlsx workbook")
        if header_row != 1:
            raise ValueError("CSV input requires header_row 1")
        with path.open(newline="", encoding="utf-8-sig") as handle:
            reader = csv.DictReader(handle)
            rows = [row for row in reader if any(text(value) for value in row.values())]
            return list(reader.fieldnames or []), rows, {
                "workbook_name": path.name,
                "selected_sheet": "",
                "selected_sheet_index": None,
                "excluded_sheet_count": 0,
                "header_row": 1,
                "first_data_row": 2,
                "completely_empty_rows": 0,
            }

    if path.suffix.lower() != ".xlsx":
        raise ValueError("input must be .xlsx or .csv")

    try:
        from openpyxl import load_workbook
    except ImportError as error:
        raise RuntimeError("openpyxl is required for .xlsx input") from error

    workbook = load_workbook(path, read_only=True, data_only=True)
    if not workbook.sheetnames:
        workbook.close()
        raise ValueError("workbook has no sheets")

    if sheet_selection == "first_sheet":
        selected_sheet = workbook.sheetnames[0]
        selected_sheet_index = 1
    elif sheet_selection == "named_sheet":
        if not sheet_name:
            workbook.close()
            raise ValueError("named_sheet selection requires sheet_name")
        if sheet_name not in workbook.sheetnames:
            workbook.close()
            raise ValueError(f"mapped sheet does not exist: {sheet_name}")
        selected_sheet = sheet_name
        selected_sheet_index = workbook.sheetnames.index(sheet_name) + 1
    else:
        workbook.close()
        raise ValueError("sheet_selection must be first_sheet or named_sheet")

    worksheet = workbook[selected_sheet]
    iterator = worksheet.iter_rows(values_only=True)
    try:
        for row_number, values in enumerate(iterator, start=1):
            if row_number == header_row:
                headers = [text(value) for value in values]
                break
        else:
            raise StopIteration
    except StopIteration:
        workbook.close()
        raise ValueError(f"header_row {header_row} does not exist in selected sheet: {selected_sheet}") from None
    data_rows = list(iterator)
    while data_rows and not any(text(value) for value in data_rows[-1]):
        data_rows.pop()
    completely_empty_rows = sum(
        not any(text(value) for value in values)
        for values in data_rows
    )
    rows = [
        {
            "__source_row_number": row_number,
            **dict(zip(headers, values)),
        }
        for row_number, values in enumerate(data_rows, start=header_row + 1)
        if any(text(value) for value in values)
    ]
    workbook_metadata = {
        "workbook_name": path.name,
        "selected_sheet": selected_sheet,
        "selected_sheet_index": selected_sheet_index,
        "excluded_sheet_count": len(workbook.sheetnames) - 1,
        "header_row": header_row,
        "first_data_row": header_row + 1,
        "completely_empty_rows": completely_empty_rows,
    }
    workbook.close()
    return headers, rows, workbook_metadata


def load_csv(path: Path) -> list[dict[str, str]]:
    with path.open(newline="", encoding="utf-8-sig") as handle:
        return list(csv.DictReader(handle))


def build_master_indexes(repo_root: Path):
    courses = load_csv(repo_root / OCC_COURSES)
    professors = load_csv(repo_root / OCC_PROFESSORS)
    relationships = load_csv(repo_root / OCC_RELATIONSHIPS)

    course_index: dict[str, set[str]] = defaultdict(set)
    for course in courses:
        canonical = compact(f"{course['subject']} {course['course_number']}")
        course_index[normalize_course(canonical)].add(canonical)

    professor_index: dict[str, set[str]] = defaultdict(set)
    for professor in professors:
        canonical = compact(professor["instructor_name"])
        for variant in professor_variants(canonical):
            professor_index[variant].add(canonical)

    relationship_index = {
        (
            normalize_course(f"{row['subject']} {row['course_number']}"),
            normalize_professor(row["instructor_name"]),
        )
        for row in relationships
    }
    return course_index, professor_index, relationship_index


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--input", required=True, type=Path)
    parser.add_argument("--mapping", required=True, type=Path)
    parser.add_argument("--output-dir", required=True, type=Path)
    parser.add_argument("--repo-root", type=Path, default=Path(__file__).resolve().parents[1])
    args = parser.parse_args()

    mapping = json.loads(args.mapping.read_text(encoding="utf-8"))
    source_type = text(mapping.get("source_type"))
    if not re.fullmatch(r"[a-z0-9]+(?:_[a-z0-9]+)*", source_type):
        raise ValueError("source_type must be a non-personal lowercase snake_case identifier")

    columns = {
        key: text(value) if value is not None else None
        for key, value in (mapping.get("columns") or {}).items()
    }
    required_keys = {
        "year_taken", "semester", "course_code", "professor_name",
        "professor_quality", "easy_a", "course_quality", "recommendation",
        "class_format", "content",
    }
    if set(columns) != required_keys:
        raise ValueError("mapping columns must contain exactly the documented Review v2 fields")

    sheet_selection = text(mapping.get("sheet_selection")) or "named_sheet"
    header_row = mapping.get("header_row", 1)
    if not isinstance(header_row, int) or isinstance(header_row, bool) or header_row < 1:
        raise ValueError("header_row must be a positive integer")
    headers, source_rows, workbook_metadata = read_rows(
        args.input,
        sheet_selection=sheet_selection,
        sheet_name=text(mapping.get("sheet_name")),
        header_row=header_row,
    )
    mapped_headers = {text(value) for value in columns.values() if value}
    missing_headers = sorted(mapped_headers - set(headers))
    if missing_headers:
        raise ValueError(f"mapped headers are missing: {missing_headers}")

    sensitive_headers = sorted(header for header in headers if SENSITIVE_HEADER.search(header))
    unmapped_headers = sorted({header for header in headers if header and header not in mapped_headers})
    if mapped_headers.intersection(sensitive_headers):
        raise ValueError("a sensitive column may not be mapped")

    semester_values = {compact(key).casefold(): value for key, value in (mapping.get("semester_values") or {}).items()}
    class_format_values = {compact(key).casefold(): value for key, value in (mapping.get("class_format_values") or {}).items()}
    if not set(semester_values.values()).issubset(SEMESTERS):
        raise ValueError("semester mapping contains an unsupported canonical value")
    if not set(class_format_values.values()).issubset(CLASS_FORMATS):
        raise ValueError("class format mapping contains an unsupported canonical value")

    course_index, professor_index, relationship_index = build_master_indexes(args.repo_root)
    preview: list[dict[str, str]] = []
    missing_recommendation_count = 0
    missing_comment_count = 0

    professor_matched_count = 0
    course_matched_count = 0
    relationship_matched_count = 0

    for source in source_rows:
        row_number = int(source["__source_row_number"])
        raw_course = text(source.get(columns["course_code"]))
        raw_professor = text(source.get(columns["professor_name"]))
        content = text(source.get(columns["content"]))
        professor_quality, professor_quality_ok = parse_rating(source.get(columns["professor_quality"]), required=True)
        easy_a, easy_a_ok = parse_rating(source.get(columns["easy_a"]), required=True)
        course_quality, course_quality_ok = parse_rating(source.get(columns["course_quality"]), required=True)
        recommendation_header = columns.get("recommendation")
        raw_recommendation = text(source.get(recommendation_header)) if recommendation_header else ""
        recommendation, recommendation_ok = parse_rating(
            raw_recommendation,
            required=False,
        )
        year_taken, year_ok = parse_year(source.get(columns["year_taken"]))
        semester = semester_values.get(compact(source.get(columns["semester"])).casefold(), "")
        class_format = class_format_values.get(compact(source.get(columns["class_format"])).casefold(), "")
        if not raw_recommendation:
            missing_recommendation_count += 1
        if not content:
            missing_comment_count += 1

        normalized_course = normalize_course(raw_course)
        course_candidates = course_index.get(normalized_course, set())
        professor_candidates: set[str] = set()
        for variant in professor_variants(raw_professor):
            professor_candidates.update(professor_index.get(variant, set()))

        canonical_course = next(iter(course_candidates)) if len(course_candidates) == 1 else ""
        canonical_professor = next(iter(professor_candidates)) if len(professor_candidates) == 1 else ""
        if canonical_course:
            course_matched_count += 1
        if canonical_professor:
            professor_matched_count += 1
        relationship_matched = bool(
            canonical_course
            and canonical_professor
            and (normalize_course(canonical_course), normalize_professor(canonical_professor)) in relationship_index
        )
        if relationship_matched:
            relationship_matched_count += 1
        pii = contains_personal_data([content])
        issues: list[str] = []
        if not all((professor_quality_ok, easy_a_ok, course_quality_ok)):
            issues.append("invalid_required_rating")
        if not recommendation_ok:
            issues.append("invalid_recommendation")
        if not year_ok:
            issues.append("invalid_year")
        if not semester:
            issues.append("invalid_semester")
        if not class_format:
            issues.append("invalid_class_format")
        if pii:
            status = "personal_data"
            content = ""
        elif issues:
            status = "invalid"
        elif len(course_candidates) == 0 or len(professor_candidates) == 0:
            status = "not_found"
            if len(course_candidates) == 0:
                issues.append("course_not_found")
            if len(professor_candidates) == 0:
                issues.append("professor_not_found")
        elif len(course_candidates) > 1 or len(professor_candidates) > 1:
            status = "ambiguous"
            if len(course_candidates) > 1:
                issues.append("course_ambiguous")
            if len(professor_candidates) > 1:
                issues.append("professor_ambiguous")
        elif not relationship_matched:
            status = "invalid_relationship"
            issues.append("invalid_professor_course_relationship")
        else:
            status = "matched"

        key_material = "|".join((
            source_type,
            normalize_course(canonical_course or raw_course),
            normalize_professor(canonical_professor or raw_professor),
            year_taken,
            semester,
            hashlib.sha256(content.encode("utf-8")).hexdigest(),
        ))
        source_row_key = hashlib.sha256(key_material.encode("utf-8")).hexdigest()

        preview.append({
            "source_type": source_type,
            "source_row_key": source_row_key,
            "source_row_number": str(row_number),
            "status": status,
            "issues": ";".join(issues),
            "course_code": canonical_course,
            "professor_name": canonical_professor,
            "professor_quality": professor_quality,
            "easy_a": easy_a,
            "course_quality": course_quality,
            "recommendation": recommendation,
            "class_format": class_format,
            "year_taken": year_taken,
            "semester": semester,
            "content": content,
        })

    duplicate_keys = {key for key, count in Counter(row["source_row_key"] for row in preview).items() if count > 1}
    for row in preview:
        if row["source_row_key"] in duplicate_keys and row["status"] == "matched":
            row["status"] = "duplicate"

    candidates = [row for row in preview if row["status"] == "matched"]
    status_counts = Counter(row["status"] for row in preview)
    valid_rows = len(preview) - status_counts["invalid"] - status_counts["personal_data"]
    report = {
        "source_type": source_type,
        "input_sha256": hashlib.sha256(args.input.read_bytes()).hexdigest(),
        "workbook_name": workbook_metadata["workbook_name"],
        "sheet_selection": sheet_selection,
        "selected_sheet": workbook_metadata["selected_sheet"],
        "selected_sheet_index": workbook_metadata["selected_sheet_index"],
        "excluded_sheet_count": workbook_metadata["excluded_sheet_count"],
        "header_row": workbook_metadata["header_row"],
        "sheet_name": workbook_metadata["selected_sheet"],
        "raw_rows": len(preview),
        "completely_empty_rows": workbook_metadata["completely_empty_rows"],
        "valid_rows": valid_rows,
        "rows": len(preview),
        "columns": headers,
        "mapped_columns": sorted(mapped_headers),
        "unmapped_columns": unmapped_headers,
        "excluded_sensitive_columns": sensitive_headers,
        "matched": status_counts["matched"],
        "professor_matched": professor_matched_count,
        "course_matched": course_matched_count,
        "relationship_matched": relationship_matched_count,
        "master_not_found": status_counts["not_found"],
        "ambiguous": status_counts["ambiguous"],
        "not_found": status_counts["not_found"],
        "duplicates": status_counts["duplicate"],
        "invalid": status_counts["invalid"],
        "invalid_rating": sum(
            bool({"invalid_required_rating", "invalid_recommendation"}.intersection(row["issues"].split(";")))
            for row in preview
        ),
        "invalid_required_rating": sum("invalid_required_rating" in row["issues"].split(";") for row in preview),
        "personal_data": status_counts["personal_data"],
        "pii_excluded": status_counts["personal_data"],
        "invalid_relationship": status_counts["invalid_relationship"],
        "missing_recommendation": missing_recommendation_count,
        "missing_comment": missing_comment_count,
        "imported_candidate_rows": len(candidates),
        "final_import_candidates": len(candidates),
    }

    args.output_dir.mkdir(parents=True, exist_ok=True)
    for filename, rows in (("preview.csv", preview), ("import_candidates.csv", candidates)):
        with (args.output_dir / filename).open("w", newline="", encoding="utf-8") as handle:
            writer = csv.DictWriter(handle, fieldnames=OUTPUT_FIELDS)
            writer.writeheader()
            writer.writerows(rows)
    (args.output_dir / "report.json").write_text(json.dumps(report, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    print(json.dumps({key: value for key, value in report.items() if key not in {"columns", "mapped_columns"}}, ensure_ascii=False))


if __name__ == "__main__":
    main()
